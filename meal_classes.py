import openpyxl as xl
import math

class Ingredient(object):
    def __init__(self, item_columns, headers, row=0):
        self.row = row + 1
        self.item_info = item_columns
        self.headers = headers
        self.category = ["Frozen / Other", ""]
        for i, header in enumerate(headers):
            if header == "Ingredient":
                self.name = [item_columns[i], header]
            if header == "Category":
                self.category = [item_columns[i], header]
            if header == "Amt":
                self.amount = [item_columns[i], header]
            if header == "Unit" or header == "Cooking Unit":
                self.unit = [item_columns[i], header]
            if header == "Price":
                self.price = [item_columns[i], header]
            if header == "Store":
                self.store = [item_columns[i], header]



class Recipe(object):
    def __init__(self, name, ingredients):
        self.row = 0
        self.name = name
        self.ingredients = ingredients


def get_headers(worksheet):
    column = worksheet.max_column
    headers = []
    for i in range(column):
        header = worksheet.cell(row=1, column=i+1).value
        headers.append(header)

    return headers


def get_ingredients(worksheet):

    row = worksheet.max_row
    column = worksheet.max_column
    headers = get_headers(worksheet)
    ingredients = []

    for i in range(row):
        ing_columns = []
        for j in range(column):
            ing_columns.append(worksheet.cell(row=i+1, column=j+1).value)

        ingredient = Ingredient(ing_columns, headers, i)
        ingredients.append(ingredient)

    return ingredients


def get_recipes(worksheet):
    recipes = []
    row = worksheet.max_row
    for i in range(1, row+1):
        if worksheet.cell(row=i+1, column=1).value is not None:
            recipes.append(worksheet.cell(row=i+1, column=1).value)

    return recipes


def get_extra(worksheet, header):
    column = worksheet.max_column
    col = 0
    extra_info = []
    for i in range(column):
        if header == worksheet.cell(row=1, column=i+1).value:
            col = i+1
            break
    if col != 0:
        for i in range(worksheet.max_row-1):
            value = worksheet.cell(row=i+2, column=col).value
            if value is None:
                continue
            extra_info.append(value)
    else:
        print("Error: Could not find column header '%s'" % header)
        return False
    return extra_info


def conversion(rec_ing, store_ing, worksheet=None):
    # from recipe_ingredients to store_ingredients
    print("Store ingredient price is " + str(store_ing.price[0]))
    print("Recipe amount is " + str(rec_ing.amount[0]))
    amount = float(rec_ing.amount[0])
    store_ing.price[0] = float(store_ing.price[0])
    if rec_ing.unit[0] == "can" and store_ing.unit[0] == "can":
        return amount, amount * store_ing.price[0], store_ing.unit[0]
    elif rec_ing.unit[0] == "whole" and store_ing.unit[0] == "whole":
        return amount, amount * store_ing.price[0], store_ing.unit[0]
    elif rec_ing.unit[0] == "cup" and store_ing.unit[0] == "can":
        amount = math.ceil(rec_ing.amount[0] * 237/400)
        return amount, amount * store_ing.price[0], store_ing.unit[0]
    elif rec_ing.unit[0] == "whole" and store_ing.unit[0] == "lbs":
        amount, price, unit = convert_to_lbs(worksheet, rec_ing, store_ing.price[0])
        return amount, price, unit
    elif rec_ing.unit[0] == "lbs" and store_ing.unit[0] == "lbs":
        return amount, amount * store_ing.price[0], store_ing.unit[0]
    else:
        return rec_ing.amount[0], store_ing.price[0], rec_ing.unit[0]

# 237 mL/cup
# 15 mL/tbsp
# 5 mL/tsp
# 3750 mL/gal

def convert_to_lbs(worksheet, rec_ing, price):
    ws1 = worksheet
    row = ws1.max_row
    for i in range(1, ws1.max_row+1):
        if rec_ing.name[0] == ws1.cell(row=i, column=1).value:
            row = i
            break
    amount = float(rec_ing.amount[0]) * ws1.cell(row=row, column=3).value
    price = amount * price
    return amount, "lbs", price


