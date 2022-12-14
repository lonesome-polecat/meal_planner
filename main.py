import openpyxl as xl
import meal_classes as mc
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
from PIL import ImageTk, Image


ents = []
filename = None
meals_wb = None
add_more = True
all_ingredients = []
store = None
recipe = None
recipe_name = None
meal_days: int = 0
all_meals = []


class IngWindow(object):
    def __init__(self):
        self.fields = mc.get_headers(meals_wb["Ingredients"])

    def main_window(self):
        self.root = tk.Tk()
        self.root.title("Ingredient")
        self.menu_bar = tk.Menu(self.root)
        self.back_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.back_menu.add_command(label=f"Main menu", command=self.go_back)
        self.menu_bar.add_cascade(label="Back", menu=self.back_menu)
        self.root.config(menu=self.menu_bar)
        self.frame = tk.LabelFrame(self.root)
        self.frame.pack()
        self.entries = self.makeform()
        options = mc.get_extra(meals_wb["Ingredients"], "Ingredient")
        self.entries[0].bind('<KeyRelease>', (lambda event: self.search(self.entries[0], options)))
        self.entries[0].focus_set()
        b1 = tk.Button(self.frame, text="Add Ingredient", command=lambda: [self.fetch()])
        b1.grid(row=5, column=1, padx=10, pady=10)
        self.root.bind('<Return>', (lambda event: self.fetch()))
        b2 = tk.Button(self.frame, text="Finish", command=lambda: [self.go_back()])
        b2.grid(row=5, column=0, padx=10, pady=10)
        self.root.mainloop()

    def makeform(self):
        entries = []
        for i, field in enumerate(self.fields):
            if field == "Category":
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                options = mc.get_extra(meals_wb["Extra"], "Category")
                ent = tk.StringVar(value=options[0])
                drop = ttk.OptionMenu(self.frame, ent, options[0], *options)
                drop.grid(row=i, column=1, padx=5, pady=8)
            elif field == "Unit":
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                options = mc.get_extra(meals_wb["Extra"], "Unit")
                ent = tk.StringVar(value=options[0])
                drop = ttk.OptionMenu(self.frame, ent, options[0], *options)
                drop.grid(row=i, column=1, padx=5, pady=8)
            elif field == "Store":
                continue
            elif field == "Ingredient":
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                options = mc.get_extra(meals_wb["Ingredients"], "Ingredient")
                ent = ttk.Combobox(self.frame, values=options)
                ent.grid(row=i, column=1, padx=5, pady=8)
            else:
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                ent = tk.Entry(self.frame)
                ent.grid(row=i, column=1, padx=5, pady=8)
            entries.append(ent)
        return entries

    def fetch(self):
        self.item_info = []
        self.headers = []
        for i, entry in enumerate(self.entries):
            self.item_info.append(entry.get().lower())
            self.headers.append(self.fields[i])

        for i, entry in enumerate(self.entries):
            if type(entry) == tk.Entry:
                entry.delete(0, "end")

        self.entries[0].focus_set()

        self.item_info.append(store)
        ing = mc.Ingredient(self.item_info, self.headers)
        all_ingredients.append(ing)

    def search(self, ent, options):
        value = ent.get()
        if value == '':
            ent['value'] = options
        else:
            data = []
            for item in options:
                if str(value).lower() in str(item).lower():
                    data.append(item)
                    ent['values'] = data

    def go_back(self):
        self.root.destroy()
        start_window()


class RecipeWindow(object):
    def __init__(self, name):
        self.name = name
        self.fields = mc.get_headers(meals_wb["Recipes"])
        print(self.fields)
        self.all_ingredients = []

    def main_window(self):
        self.root = tk.Tk()
        self.root.title("Recipe")
        self.menu_bar = tk.Menu(self.root)
        self.back_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.back_menu.add_command(label=f"Main menu", command=self.go_back)
        self.menu_bar.add_cascade(label="Back", menu=self.back_menu)
        self.root.config(menu=self.menu_bar)
        self.frame = tk.LabelFrame(self.root)
        self.frame.pack()
        self.entries = self.makeform()
        options = self.entries[0]['values']
        self.entries[0].bind('<KeyRelease>', (lambda event: self.search(self.entries[0], options)))
        self.entries[0].focus_set()
        b1 = tk.Button(self.frame, text="Add Ingredient", command=lambda: [self.fetch()])
        b1.grid(row=5, column=1, padx=10, pady=10)
        self.root.bind("<Return>", (lambda event: self.fetch()))
        b2 = tk.Button(self.frame, text="Finish", command=lambda: [self.finish(), self.go_back()])
        b2.grid(row=5, column=0, padx=10, pady=10)
        self.root.mainloop()

    def makeform(self):
        # Make 'Ingredient' field a combobox with values from 'ingredients' sheet
        entries = []
        for i, field in enumerate(self.fields):
            if field == "Unit":
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                options = mc.get_extra(meals_wb["Extra"], "Cooking Unit")
                ent = tk.StringVar(value=options[0])
                drop = ttk.OptionMenu(self.frame, ent, options[0], *options)
                drop.grid(row=i, column=1, padx=5, pady=8)
            elif field == "Recipe":
                continue
            elif field == "Ingredient":
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                options = mc.get_extra(meals_wb["Ingredients"], "Ingredient")
                ent = ttk.Combobox(self.frame, values=options)
                ent.grid(row=i, column=1, padx=5, pady=8)
            else:
                label = tk.Label(self.frame, text=field)
                label.grid(row=i, column=0, padx=5, pady=8)
                ent = tk.Entry(self.frame)
                ent.grid(row=i, column=1, padx=5, pady=8)
            entries.append(ent)
        return entries

    def fetch(self):
        self.item_info = []
        self.headers = []
        for i, entry in enumerate(self.entries):
            self.item_info.append(entry.get().lower())
            self.headers.append(self.fields[i+1])

        for i, entry in enumerate(self.entries):
            if type(entry) == tk.Entry or type(entry) == ttk.Combobox:
                entry.delete(0, "end")

        self.entries[0].focus_set()

        ing = mc.Ingredient(self.item_info, self.headers)
        print(f"{ing.name[0]} {str(ing.amount[0])} {ing.unit[0]}")
        self.all_ingredients.append(ing)

    def search(self, ent, options):
        value = ent.get()
        if value == '':
            ent['value'] = options
        else:
            data = []
            for item in options:
                if str(value).lower() in str(item).lower():
                    data.append(item)
                    ent['values'] = data

    def finish(self):
        self.root.destroy()
        upload_recipe()
        print("Uploaded recipe")
        start_window()

    def go_back(self):
        self.root.destroy()
        start_window()


class MealPlanWindow(object):
    def __init__(self):
        self.fields = mc.get_headers(meals_wb["Schedule"])
        self.all_recipes = []
        self.all_ingredients = []
        self.temp_meal_list = []

    def main_window(self, edit_window=None):
        if edit_window is not None:
            del edit_window
            self.all_ingredients.clear()
            self.all_recipes.clear()
        self.root = tk.Tk()
        self.menu_bar = tk.Menu(self.root)
        self.back_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.back_menu.add_command(label=f"Main menu", command=self.go_back)
        self.menu_bar.add_cascade(label="Back", menu=self.back_menu)
        self.root.config(menu=self.menu_bar)
        frame = tk.LabelFrame(self.root)
        frame.pack(side=tk.TOP, padx=10, pady=10)
        label = tk.Label(frame, text="Select meals/sides you would like to make")
        label.pack(side=tk.TOP, padx=10, pady=10)
        options = mc.get_recipes(meals_wb["Recipes"])
        ent = ttk.Combobox(frame, values=options)
        ent.pack(side=tk.TOP, padx=20, pady=10, fill=tk.X)
        ent.focus_force()
        add_button_frame = tk.Frame(frame)
        add_button_frame.pack(side=tk.TOP, fill=tk.BOTH)
        self.listbox = tk.Listbox(frame)
        self.listbox.pack(side=tk.TOP, padx=20, pady=10, fill=tk.X)
        if len(self.temp_meal_list) > 0:
            for meal in self.temp_meal_list:
                self.listbox.insert("end", meal)
        remove_button_frame = tk.Frame(frame)
        remove_button_frame.pack(side=tk.TOP, fill=tk.BOTH)
        add_button = tk.Button(add_button_frame, text="Add meal", command=lambda: self.update_list(self.listbox, ent))
        add_button.pack(side=tk.RIGHT, padx=20, pady=3)
        remove_button = tk.Button(remove_button_frame, text="Remove meal", command=lambda: self.remove_from_list(self.listbox))
        remove_button.pack(side=tk.RIGHT, padx=20, pady=(3, 10))
        continue_button_frame = tk.Frame(self.root)
        continue_button_frame.pack(side=tk.BOTTOM, fill=tk.BOTH)
        b1 = tk.Button(continue_button_frame, text="Continue",
                       command=lambda: [self.fetch()])
        b1.pack(side=tk.BOTTOM, padx=10, pady=15)
        ent.bind('<Return>', (lambda event: self.update_list(self.listbox, ent)))
        # self.root.bind('<Return>', (lambda event: [fetch(ent, 3), root.destroy(), meal_plan()]))
        self.root.mainloop()
        """
        self.root = tk.Tk()
        self.root.title("Schedule")
        self.frame = tk.LabelFrame(self.root)
        self.frame.pack()
        self.entries = self.makeform()
        self.root.bind("<Return>", (lambda event: [self.fetch(), self.root.destroy()]))
        b2 = tk.Button(self.frame, text="Finish", command=lambda: [self.fetch(), self.root.destroy()])
        b2.grid(column=1, padx=10, pady=10)
        self.root.mainloop()
        """

    def makeform(self):
        entries = []
        for i in range(int(meal_days)):
            label = tk.Label(self.frame, text=f"Meal {i+1}")
            label.grid(row=i, column=0, padx=5, pady=8)
            options = mc.get_recipes(meals_wb["Recipes"])
            ent = ttk.Combobox(self.frame, values=options)
            ent.grid(row=i, column=1, padx=5, pady=8)
            ent.bind('<KeyRelease>', (lambda event: self.search(ent, options)))
            entries.append(ent)
        return entries

    def remove_from_list(self, list):
        field = list.selection_get()
        index = list.get(0, "end").index(field)
        self.temp_meal_list.pop(index)
        list.delete(index)
        print(self.temp_meal_list)

    def update_list(self, list, entry):
        if entry.get() == "":
            return None
        list.insert("end", entry.get())
        self.temp_meal_list.append(entry.get())
        entry.delete(0, "end")
        print(self.temp_meal_list)

    def fetch(self):
        """Adds meals to 'Schedule' and moves to EditWindow"""
        for meal in self.listbox.get(0, "end"):
            self.all_recipes.append(meal)

        global all_meals
        all_meals = self.all_recipes
        ws1 = meals_wb["Recipes"]
        headers = mc.get_headers(ws1)
        recipe_count = 0
        for i in range(1, ws1.max_row+1):
            cell = ws1.cell(row=i+1, column=1).value
            print(cell)
            if cell is not None:
                if cell in self.all_recipes:
                    j = i
                    while ws1.cell(row=j+1, column=2).value != None:
                        row = j+1
                        name = ws1.cell(row=row, column=headers.index("Ingredient")+1).value
                        amount = ws1.cell(row=row, column=headers.index("Amt")+1).value
                        unit = ws1.cell(row=row, column=headers.index("Unit")+1).value
                        item_info = [name, amount, unit]
                        my_ing = mc.Ingredient(item_info, headers[1:])
                        print(my_ing.name[0])
                        self.all_ingredients.append(my_ing)
                        j += 1
                    recipe_count += 1
                    if recipe_count == len(self.all_recipes):
                        break

        self.root.destroy()
        upload_meals(self)

    def search(self, ent, options):
        value = ent.get()
        if value == '':
            ent['value'] = options
        else:
            data = []
            for item in options:
                if str(value).lower() in str(item).lower():
                    data.append(item)
                    ent['values'] = data

    def go_back(self):
        self.root.destroy()
        start_window()


class Category_Frame(object):
    def __init__(self, name, list):
        self.name = name
        self.column_index = None
        self.frame = None
        self.list = list
        self.buttons = []


class EditWindow(object):
    def __init__(self):
        self.ws = meals_wb["ShoppingList"]
        self.shopping_dict = {}
        self.headers = []
        for i in range(self.ws.max_column):
            header = self.ws.cell(row=1, column=i+1).value
            self.headers.append(header)
            if header == "Amt":
                continue
            values = []
            for j in range(self.ws.max_row):
                value = self.ws.cell(row=j+2, column=i+1).value
                if value == "":
                    break
                values.append(value)
            self.shopping_dict.update({header: values})
        self.frames = []

    def main_window(self):
        self.root = tk.Tk()
        self.menu_bar = tk.Menu(self.root)
        self.back_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.back_menu.add_command(label=f"Meals", command=self.back_to_meals)
        self.menu_bar.add_cascade(label="Back", menu=self.back_menu)
        self.root.config(menu=self.menu_bar)
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(side=tk.TOP)
        self.button_frame = tk.Frame(self.root)
        self.button_frame.pack(side=tk.TOP)
        self.finish_button = tk.Button(self.button_frame, text="Finish", command=lambda: self.finish())
        self.finish_button.pack(side=tk.RIGHT, padx=10, pady=10)
        self.add_to_frame()
        self.root.mainloop()

    def add_to_frame(self):
        for key, values in self.shopping_dict.items():
            list = self.get_list_frame(key, values)
            category_frame = Category_Frame(key, list)
            self.frames.append(category_frame)

    def get_list_frame(self, header, values):
        frame = tk.LabelFrame(self.main_frame, text=f"{header}")
        frame.pack(side=tk.LEFT, padx=10, pady=10)
        list = tk.Listbox(frame)
        list.pack(padx=10, pady=10)
        for i in range(len(values)):
            list.insert("end", values[i])
        button_frame = tk.Frame(frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.BOTH)
        add_button = tk.Button(button_frame, text="Add", command=lambda list=list: self.add_item(list), anchor="e")
        add_button.grid(row=0, column=0, pady=5)
        remove_button = tk.Button(button_frame, text="Remove", anchor="w",
                                  command=lambda list=list: self.remove_item(list, list.selection_get()))
        remove_button.grid(row=0, column=1, pady=5)
        return list

    def remove_item(self, list, field):
        index = list.get(0, "end").index(field)
        list.delete(index)

    def add_item(self, list):
        self.add_window = tk.Toplevel(self.root)
        label = tk.Label(self.add_window, text="Add item")
        label.grid(row=0, column=0, padx=10, pady=10)
        entry = tk.Entry(self.add_window)
        entry.grid(row=0, column=1, padx=10, pady=10)
        button = tk.Button(self.add_window, text="Add", command=lambda list=list: self.update_list(list, entry))
        button.grid(row=1, column=1, padx=10, pady=10)

    def update_list(self, list, entry):
        list.insert("end", entry.get())
        self.add_window.destroy()

    def finish(self):
        ws1 = meals_wb['ShoppingList']
        ws1.delete_rows(2, ws1.max_row)
        for frame in self.frames:
            index = self.headers.index(frame.name)
            items = frame.list.get(0, "end")
            for i in range(len(items)):
                self.ws.cell(row=i+2, column=index+1).value = items[i]
        meals_wb.save(filename)
        print("Shopping list has been updated")
        self.root.destroy()

    def back_to_meals(self):
        self.root.destroy()
        meals.main_window(self)


def fetch(ent, index=0):
    global store
    global recipe_name
    global meal_days
    if index == 1:
        store = ent.get()
    elif index == 2:
        recipe_name = ent.get()
    elif index == 3:
        meal_days = ent.get()


def store_prompt():
    root = tk.Tk()
    frame = tk.Frame(root)
    frame.pack(side=tk.TOP, padx=5, pady=5)
    label = tk.Label(frame, text="Enter the name of the store you shop at")
    label.pack(side=tk.TOP, padx=10, pady=10)
    options = mc.get_extra(meals_wb["Extra"], "Store")
    ent = ttk.Combobox(frame, values=options)
    ent.pack(side=tk.TOP, padx=10, pady=10)
    ent.focus_force()
    b1 = tk.Button(frame, text="Continue", command=lambda: [fetch(ent, 1), root.destroy(), add_ingredients()])
    b1.pack(side=tk.BOTTOM, padx=10, pady=10)
    root.bind('<Return>', (lambda event: [fetch(ent, 1), root.destroy(), add_ingredients()]))
# In class windows have focus_set()


def recipe_name_prompt():
    root = tk.Tk()
    frame = tk.Frame(root)
    frame.pack(side=tk.TOP, padx=5, pady=5)
    label = tk.Label(frame, text="Enter the name of the recipe")
    label.pack(side=tk.TOP, padx=10, pady=10)
    ent = tk.Entry(frame)
    ent.pack(side=tk.TOP, padx=10, pady=10)
    ent.focus_force()
    b1 = tk.Button(frame, text="Continue", command=lambda: [fetch(ent, 2), root.destroy(), add_recipe()])
    b1.pack(side=tk.BOTTOM, padx=10, pady=10)
    root.bind('<Return>', (lambda event: [fetch(ent, 2), root.destroy(), add_recipe()]))


def remove_from_list(list):
    field = list.selection_get()
    index = list.get(0, "end").index(field)
    list.delete(index)


def update_list(list, entry):
    list.insert("end", entry.get())
    entry.delete(0, "end")


def add_recipe():
    global recipe
    recipe = RecipeWindow(recipe_name)
    recipe.main_window()


def add_ingredients():
    ing_window = IngWindow()
    ing_window.main_window()
    upload_ingredients()


def upload_ingredients():
    ing_ws = meals_wb["Ingredients"]
    headers = mc.get_headers(ing_ws)
    preloaded_ings = mc.get_extra(ing_ws, "Ingredient")
    for i, ingredient in enumerate(all_ingredients):
        row = ing_ws.max_row
        if ingredient.name[0] in preloaded_ings:
            row = preloaded_ings.index(ingredient.name[0])+1
        for j, header in enumerate(headers):
            ing_ws.cell(row=row+1, column=j+1).value = ingredient.item_info[j]

    meals_wb.save(filename)


def upload_recipe():
    global recipe
    ws1 = meals_wb["Recipes"]
    ws2 = meals_wb["Ingredients"]
    ingredients = mc.get_extra(meals_wb["Ingredients"], "Ingredient")
    headers = mc.get_headers(ws1)
    row = ws1.max_row + 1
    ws2_row = ws2.max_row + 1
    if row > 3:
        row += 1
    ws1.cell(row=row, column=headers.index("Recipe")+1).value = recipe.name
    for ing in recipe.all_ingredients:
        ws1.cell(row=row, column=headers.index(ing.name[1])+1).value = ing.name[0]
        ws1.cell(row=row, column=headers.index(ing.amount[1])+1).value = ing.amount[0]
        ws1.cell(row=row, column=headers.index(ing.unit[1])+1).value = ing.unit[0]
        if ing.name[0].lower() not in ingredients:
            ws2.cell(row=ws2_row, column=1).value = ing.name[0]
            ws2.cell(row=ws2_row, column=2).value = "other"
            ws2.cell(row=ws2_row, column=3).value = 0
            ws2.cell(row=ws2_row, column=4).value = "none"
            ws2.cell(row=ws2_row, column=5).value = 1.00
            ws2_row += 1
        row += 1

    meals_wb.save(filename)


def upload_meals(meals):
    print("Uploading...")
    rec_ings = meals.all_ingredients
    print(f"There are {len(rec_ings)} ingredients total")
    names = []
    for i, ing in enumerate(rec_ings):
        if ing.name[0].lower() in names:
            rec_ings[names.index(ing.name[0])].amount[0] = float(rec_ings[names.index(ing.name[0])].amount[0]) + float(ing.amount[0])
            rec_ings.pop(i)
            print("Removed duplicate ingredient")
        names.append(ing.name[0].lower())

    all_ings = mc.get_ingredients(meals_wb['Ingredients'])
    ws1 = meals_wb['ShoppingList']
    headers = mc.get_headers(ws1)
    ws1.delete_rows(2, ws1.max_row)
    prices = []
    column = ""
    # for i in range(1, ws1.max_row):

    row = ws1.max_row + 1
    for i, rec_ing in enumerate(rec_ings):
        isFound = False
        for j, ingredient in enumerate(all_ings):
            ingredient.category[0] = ingredient.category[0].capitalize()
            if ingredient.name[0].lower() == rec_ing.name[0].lower():
                #print(store_ing.category[0])
                amount, price, unit = mc.conversion(rec_ing, ingredient, meals_wb['Weights'])
                if ingredient.category[0] == "Grains" or ingredient.category[0] == "Can" or \
                        ingredient.category[0] == "Bottle" or ingredient.category[0] == "Spice":
                    ingredient.category[0] = "Main aisles"
                elif ingredient.category[0].lower() == "frozen" or ingredient.category[0].lower() == "other":
                    ingredient.category[0] = "Frozen / Other"
                column = headers.index(ingredient.category[0])+1
                for k in range(1, row+1):
                    if ws1.cell(row=k, column=column).value is None:
                        temp_row = k
                        break
                ws1.cell(row=temp_row, column=column).value = ingredient.name[0]
                ws1.cell(row=temp_row, column=column+1).value = f"{str(amount)} {unit}"
                #prices.append(price)
                isFound = True
        if not isFound:
            print("I did not find the ingredient")
            column = headers.index(rec_ing.category[0])+1
            for k in range(1, row+1):
                if ws1.cell(row=k, column=column).value is None:
                    temp_row = k
                    break
            ws1.cell(row=temp_row, column=column).value = rec_ing.name[0]
            ws1.cell(row=temp_row, column=column+1).value = f"{str(rec_ing.amount[0])} {rec_ing.unit[0]}"
        row += 1

    ws2 = meals_wb['Schedule']
    ws2.delete_rows(2, ws2.max_row)
    new_headers = mc.get_headers(ws2)
    for i, meal in enumerate(all_meals):
        ws2.cell(row=i+2, column=new_headers.index("Meal")+1).value = meal
    """
    total_price = 0
    for i in range(len(prices)):
        print(prices[i])
        prices[i] = float(prices[i])
        total_price += prices[i]
    ws2.cell(row=2, column=new_headers.index("Total")+1).value = total_price
    """

    meals_wb.save(filename)
    edit_window = EditWindow()
    edit_window.main_window()


def start_window():
    root = tk.Tk()
    root_frame = tk.Frame(root)
    root_frame.pack(side=tk.TOP, fill=tk.X)
    img = Image.open("food_image.jpg")
    ratio = 5.5
    width = 1980 / ratio
    height = 1320 / ratio
    resized = img.resize((int(width), int(height)), Image.ANTIALIAS)
    ready_img = ImageTk.PhotoImage(resized)
    image_label = tk.Label(root_frame, image=ready_img)
    image_label.grid(row=0, column=0, padx=0, pady=0)
    frame = tk.Frame(root_frame, height=40, padx=20)
    frame.grid(row=1, column=0, padx=10, pady=10)
    b1 = tk.Button(frame, text="Meal Plan", anchor="center", command=lambda: [root.destroy(), meals.main_window()])
    b2 = tk.Button(frame, text="Add Recipe", anchor="center", command=lambda: [root.destroy(), recipe_name_prompt()])
    b3 = tk.Button(frame, text="Add Ingredients", anchor="center", command=lambda: [root.destroy(), store_prompt()])
    b3.grid(row=0, column=0, padx=20, pady=5)
    b2.grid(row=0, column=1, padx=20, pady=5)
    b1.grid(row=0, column=2, padx=20, pady=5)
    root.mainloop()


def get_file_path(file_path, message):
    """
        A Tkinter Toplevel window that asks the user for input

        A popup that explains to the user that a certain .txt file
        that contains a certain file path is missing,
        and prompts the user to route it to the proper file path

        Parameters
        ---------
        file_path: list
            Empty list to hold the file path
        message: str
            The main text of the popup window
        """
    root = tk.Tk()
    frame = tk.Frame(root)
    frame.pack()
    label = tk.Label(frame, text=message)
    label.pack(side=tk.TOP, padx=10, pady=10)
    filename_entry = tk.Entry(frame)
    filename_entry.pack(side=tk.LEFT, padx=5, pady=10, fill=tk.X)
    b2 = tk.Button(frame, width=8, text="Save",
                   command=lambda: [file_path.append(filename_entry.get()), root.destroy()])
    b2.pack(side=tk.RIGHT, padx=5, pady=10)
    if "json" in message:
        b1 = tk.Button(frame, text="Browse...",
                       command=lambda: [filename_entry.delete(0, "end"), filedialogbox_dir(root, filename_entry)])
    else:
        b1 = tk.Button(frame, text="Browse...", command=lambda: [filename_entry.delete(0, "end"), filedialogbox(root, filename_entry)])
    b1.pack(side=tk.RIGHT, padx=5, pady=10)
    root.mainloop()


def filedialogbox(root: tk.Tk, ent: tk.Entry):
    """
    Asks the user to select a file and puts it into the Entry box

    Parameters
    ---------
    root: tk.Tk
        Tkinter Tk widget
    ent: tk.Entry
        Tkinter Entry box widget
    """
    root.filename = filedialog.askopenfilename(initialdir=r"C:\Users", title="Select an xl file",
                                               filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
    ent.insert(0, root.filename)


def filedialogbox_dir(root, ent):
    """
    Asks the user to select a directory (folder) and puts it into the Entry box

    Parameters
    ---------
    root: tk.Tk
        Tkinter Tk widget
    ent: tk.Entry
        Tkinter Entry box widget
    """
    root.filename = filedialog.askdirectory(initialdir=r"C:\Users", title="Select a directory")
    ent.insert(0, root.filename)


def check_and_get_file_path():
    try:
        with open(r"meal_planner_path.txt", "r+") as f:
            inv_path = f.read()
    except:
        message = "***Could not find meal_planner file***\n\nPlease specify the path to file"
        print(message)
        file_path = []
        get_file_path(file_path, message)
        inv_path = file_path[0]
        print(inv_path)

    with open(r"meal_planner_path.txt", "w") as f:
        f.write(inv_path)

    return inv_path


if __name__ == '__main__':
    filename = check_and_get_file_path()

    while True:  # repeat until the try statement succeeds
        try:
            myfile = open(filename, "r+")  # This is just to make sure that the file isn't already open
            # Because otherwise the program pulls up an error after the user has already input everything
            myfile.close()
            break  # exit the loop
        except IOError:
            input("Could not open file! Please close %s. Press Enter to retry." % filename)
            # restart the loop

    meals_wb = xl.load_workbook(filename)
    meals = MealPlanWindow()
    start_window()
